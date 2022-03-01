'''
TODO
- скачивание контрактов
- борьба с вылетами (перехват исключений на все операции с внешним миром)
- борьба с ошибками записи в Excel файл
- качественное скачивание информации о лотах на этапе 2 (улучение чтения таблицы)
- разноцветный вывод в командную строку?
- управление форматом данных в файле?
- использование последнего ИНН в новых запросах
'''

import sys
import csv
import os
import os.path
import re
import datetime
import calendar
import argparse
import json
from fnmatch import fnmatch
try:
    import openpyxl
    import requests
    from bs4 import BeautifulSoup
except ModuleNotFoundError:
    print('Ошибка загрузки модуля.\nВыполните команду:\npython -m pip install -r requirements.txt')
    raise

from bad_list import bad_list

MAIN_FILE_NAME = 'zakupki.xlsx'
headers = {
    'user-agent': 'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:97.0) Gecko/20100101 Firefox/97.0',
        }
base_url = 'https://zakupki.gov.ru'


def main():
    global script_start_time
    script_start_time = datetime.datetime.now()
    
    script_path, script_name = os.path.split(__file__)
    script_mod_time = datetime.datetime.fromtimestamp(os.path.getmtime(__file__))

    args = command_line_processing()
    
    if os.path.exists(MAIN_FILE_NAME):
        print(f'Файл {MAIN_FILE_NAME} найден, он будет дополнен')
        wb = openpyxl.open(MAIN_FILE_NAME)
    else:
        print(f'Файл {MAIN_FILE_NAME} не найден, он будет создан при закрытии программы')
        wb = openpyxl.Workbook()

    log = get_wrapper(wb, 'log', 0)

    log_entry = {
        'Date time': script_start_time,
        'Script name': script_name,
        'Script mod time': script_mod_time,
        'Command line': json.dumps(sys.argv),
        'Stage': args.stage,
        }

    if args.stage == 1:
        print('Этап 1, первичный сбор данных с сайта')
        inn_list = args.inn
        years = []
        if not inn_list:
            print('Не указано ИНН ЛПУ для сбора данных')
            print('Будут обработаны оставшиеся задачи по сбору данных (если они ещё есть)')
        else:            
            print('Обрабатываем следующие запросы:')
            for i, inn in enumerate(args.inn):
                print(f'{i+1:3d}: {inn}')
            print('Также будут обработаны оставшиеся задачи по сбору данных (если они ещё есть)')

            # Есть ИНН - нужен список лет для поиска
            current_year = script_start_time.year
            for year in args.year:
                if year < 0:
                   continue 
                if year < 100:
                    year += 2000
                if year < 2000 or year > current_year:
                    continue
                years.append(year)
            if len(years) == 0:
                print('Не указано ни одного подходящего года, добавляем текущий')
                # years.append(current_year)
                years = [current_year]
            else:
                years = sorted(set(years))
            print(f'Годы поиска: {years}')
            
            log_entry['INN list'] = json.dumps(inn_list)
            log_entry['Years'] = json.dumps(years)

        do_stage_one(wb, inn_list, years)
        
    elif args.stage == 2:
        print('Этап 2, сбор дополнительных данных по каждому лоту')
        do_stage_two(wb)
    elif args.stage == 3:
        print('Этап 3, скачивание доп информации по отмеченным лотам')
        do_stage_three(wb)

    log.append(log_entry)

    try:
        wb.save(MAIN_FILE_NAME)
    except Exception:
        alt_name = make_do_not_exists(MAIN_FILE_NAME)
        print(f'Проблема с записью в файл {MAIN_FILE_NAME}')
        print(f'Попробуем файл {alt_name}')
        wb.save(alt_name)

def command_line_processing():
    parser = argparse.ArgumentParser(
        description='Анализ торгов на сайте Zakupki.gov.ru',
        )
    parser.add_argument(
        'stage',
        metavar='STAGE',
        nargs='?',
        default=1,
        type=int,
        help='Этап работы, варианты: 1, 2 или 3',
        choices=[1, 2, 3],
        )
    parser.add_argument(
        '-i', '--inn',
        metavar='INN',
        nargs='+',
        help='ИНН (один или несколько) для сбора статистики; на этапе 1 должен быть указан. На этапах 2-3 параметр будет проигнорирован',
        dest='inn',
        default=[],
        )
    parser.add_argument(
        '-y', '--year',
        metavar='YEAR',
        nargs='+',
        type=int,
        help='Год (годы) для сбора статистики на этапе 1; если не указан: будет использован текущий год. На этапах 2-3 параметр будет проигнорирован',
        dest='year',
        default=[],
        )
    args = parser.parse_args()
    return args


class WS_wrapper:
    def __init__(self, ws):
        self.ws = ws

        if ws.max_column <= 1 and ws.max_row <= 1:
            ws['A1'] = 'ID'
            self.name_to_index = {'ID': 1}
            self.names = ['ID']
            self.new_ID = 1
        else:
            self.name_to_index = {}
            self.names = [None] * ws.max_column
            for i in range(ws.max_column):
                name = ws.cell(1, i+1).value
                self.name_to_index[name] = i+1
                self.names[i] = name
            self.new_ID = 1
            for j in range(ws.max_row - 1):
                self.new_ID = max(
                    self.new_ID,
                    1 + int(ws.cell(2+j, 1).value)
                    )

    def __len__(self):
        return self.ws.max_row-1

    def __getitem__(self, index):
        if index < 0 or index > self.ws.max_row-2:
            raise KeyError
        tmp = {}
        for i in range(self.ws.max_column):
            name = self.ws.cell(1, i+1).value
            tmp[name] = self.ws.cell(index+2, i+1).value
        return tmp

    def __setitem__(self, index, record):
        if index > self.ws.max_row-2:
            raise KeyError
        for name in record:
            if name in self.name_to_index:
                col = self.name_to_index[name]
            else:
                col = self.ws.max_column + 1
                self.name_to_index[name] = col
                self.ws.cell(1, col).value = name
            self.ws.cell(index+2, col).value = record[name]

    def append(self, record):
        index = self.ws.max_row-1
        self.ws.cell(index+2, 1).value = self.new_ID
        self.new_ID += 1
        self[index] = record


def get_wrapper(wb, sheet_name, position):
    if sheet_name in wb:
        sheet = wb[sheet_name]
    else:
        sheet = wb.create_sheet(sheet_name, position)
    wrapper = WS_wrapper(sheet)
    return wrapper


def make_do_not_exists(file_name):
    """Make name for file that do not exists

Input: `file_name` - some name for file
Output: file name (the same or modified) do not exists"""

    while os.path.exists(file_name):
        path, name = os.path.split(file_name)
        name, ext = os.path.splitext(name)
        match = re.match(r'(.*)\(([0-9]+)\)', name)
        try:
            name = f'{match[1].strip()} ({int(match[2])+1}){ext}'
        except Exception:
            name = f'{name.strip()} (1){ext}'
        file_name = os.path.join(path, name)
            
    return file_name


def do_stage_one(wb, inn_list, years):
    jobs = get_wrapper(wb, 'jobs', 1)
    current_year = script_start_time.year
    current_month = script_start_time.month

    job = {'state': 'pending'}
    for inn in inn_list:
        job['INN'] = inn
        for year in years:
            job['year'] = year
            for month in range(1, 13):
                if year == current_year and \
                   month > current_month:
                    break
                job['month'] = month
                jobs.append(job)

    lots = get_wrapper(wb, 'lots', 2)
    
    jobs_count = len(jobs)
    for i in range(jobs_count):
        job = jobs[i]

        if job['state'] == 'done':
            continue
        
        try:
            complete_the_task(lots, job)
        except Exception as e:
            print(f'Какая-то ошибка с заданием {job}, запустите программу ещё раз без аргументов для выполнения всех отложенных задач')
            print('Вот описание ошибки:')
            print(e)
            job['state'] = 'error'
        else:
            job['state'] = 'done'
            
        jobs[i] = job


def complete_the_task(lots, job):
    target_inn = str(job['INN'])
    target_year = int(job['year'])
    target_month = int(job['month'])

    print(f'Ищем аукционы для ИНН {target_inn} за год {target_year}, месяц {target_month}')
    
    extended_search = 'https://zakupki.gov.ru/epz/order/extendedsearch/results.html'
    params = {'morphology': 'on', 'sortDirection': 'false', 'recordsPerPage': '_50', 'showLotsInfoHidden': 'false',
              'sortBy': 'UPDATE_DATE', 'fz44': 'on', 'fz223': 'on', 'af': 'on', 'ca': 'on', 'pc': 'on', 'pa': 'on',
              'priceContractAdvantages44IdNameHidden': '%7B%7D', 'priceContractAdvantages94IdNameHidden': '%7B%7D',
              'currencyIdGeneral': '-1', 'selectedSubjectsIdNameHidden': '%7B%7D',
              'OrderPlacementSmallBusinessSubject': 'on', 'OrderPlacementRnpData': 'on',
              'OrderPlacementExecutionRequirement': 'on', 'orderPlacement94_0': '0', 'orderPlacement94_1': '0',
              'orderPlacement94_2': '0', 'contractPriceCurrencyId': '-1', 'budgetLevelIdNameHidden': '%7B%7D',
              'nonBudgetTypesIdNameHidden': '%7B%7D', 'searchString': target_inn}

    date1 = datetime.date(target_year, target_month, 1)
    _, m = calendar.monthrange(target_year, target_month)
    date2 = date1.replace(day=m)

    params['publishDateFrom'] = date1.strftime('%d.%m.%Y')
    params['publishDateTo'] = date2.strftime('%d.%m.%Y')

    print(f'Период от {params["publishDateFrom"]} до {params["publishDateTo"]}')

    pageNumber = 1
    while True:
        params['pageNumber'] = str(pageNumber)
        response = requests.get(
            extended_search,
            headers=headers,
            params=params,
            timeout=60,
            )
        if response.status_code != 200:
            print('Ошибка связи с сайтом. Запустите программу ещё раз')
            raise Exception(f'Request url {response.request.url} result {response.status_code}')
            
        soup = BeautifulSoup(response.text, 'html.parser')

        if pageNumber == 1:
            part = soup.find('div', 'search-results__total')
            value = part.text.strip()
            print(f'Найдено всего аукционов за период: `{value}`')
            if value.startswith('бол'):
                print('Найдено слишком много записей, попробуйте разбить диапазон времени сильнее')
                raise Exception('Too many entries')
        count = work_with_searchresult(lots, soup)
        if count == 0:
            break
        
        pageNumber += 1


def work_with_searchresult(lots, soup):
    count = 0
    for data_block in soup.find_all('div', class_='search-registry-entry-block box-shadow-search-input'):
        count += 1
        
        record = {
            'stage2': 'none',
            'stage3': 'no',
            # 'published': None,
            # 'updated': None,
            # 'last_date': None,
            }
        
        part = data_block.find('div', class_='registry-entry__header-top__title')

        values = [stripped for e in part.text.split(sep='\n') if (stripped:=e.strip())]
        assert len(values) == 2
        record['fz'] = values[0]
        record['subtype'] = values[1]
        
        part = data_block.find('div', class_='registry-entry__header-mid__number')
        a = part.find('a')
        if record['fz'] == '44-ФЗ':
            record['link'] = base_url + a['href']
        else:
            record['link'] = a['href']
        
        values = a.text.strip().split()
        assert len(values) == 2
        assert values[0] == '№'
        record['number'] = '"' + values[1] + '"'

        part = data_block.find('div', class_='registry-entry__header-mid__title')
        record['stage'] = part.text.strip()

        body = data_block.find('div', class_='registry-entry__body')

        part = body.find('div', class_='registry-entry__body-value')
        name = part.text.strip()
        record['name'] = name.replace('\n', ' ').replace('\r', ' ').replace('  ', ' ')
        low_name = record['name'].lower()

        bad_flag = False
        for bad in bad_list:
            if re.search(bad, low_name):
                bad_flag = True
                break
        if bad_flag:
            # print(f'BAD {record["name"]=}')
            continue

        part = body.find('div', class_='registry-entry__body-href')
        a = part.find('a')
        record['agency_link'] = base_url + a['href']
        record['agency'] = a.text.strip()

        right = data_block.find('div', class_='col col d-flex flex-column registry-entry__right-block b-left')

        part = right.find('div', class_='price-block__value')
        value = part.text.strip().replace('₽', '').replace(chr(160), '').replace(',', '.')
        record['price'] = float(value)

        record['published'] = None
        record['updated'] = None
        record['last_date'] = None
        
        datas = right.find('div', class_='data-block mt-auto')
        for d in datas.find_all('div', class_='data-block__title'):
            v = d.find_next_sibling('div', class_='data-block__value')
            description = d.text.strip()
            value = datetime.datetime.strptime(v.text, '%d.%m.%Y').date()
            if description == 'Размещено':
                record['published'] = value
            elif description == 'Обновлено':
                record['updated'] = value
            elif description == 'Окончание подачи заявок':
                record['last_date'] = value
            else:
                print(f'Some strange value {description=}')
                raise NotImplementedError
                
        print(f'{count:3d}: {record["name"]}')

        lots.append(record)
        
    return count

def do_stage_three(wb):
    lots = get_wrapper(wb, 'lots', 2)
    lots_count = len(lots)
    if lots_count <= 0:
        print('В файле не найдено данных о торгах')
        return

    # common_info = 'https://zakupki.gov.ru/epz/order/notice/ea44/view/common-info.html'
    # supp_base = 'https://zakupki.gov.ru/epz/order/notice/ea44/view/supplier-results.html'
    documents = 'https://zakupki.gov.ru/epz/order/notice/ea44/view/documents.html'

    for index in range(lots_count):
        record = lots[index]
        if record['stage3'] in ['no', 'done']:
            continue

        lot_tag = f'Лот {record["ID"]:03n} {record["price"]*1e-6:.3f} М руб'
        if not os.path.isdir(lot_tag):
            os.mkdir(lot_tag)

        print(lot_tag)

        number = record['number'].replace('"', '')
        params = {'regNumber': number}

        try:
            response = requests.get(
            documents,
            headers=headers,
            params=params,
            timeout=60)
        except Exception as e:
            print(f'Какая-то ошибка с получением информации с сайта, запустите программу ещё раз на третий этап')
            print('Вот описание ошибки:')
            print(e)
            continue

        if response.status_code != 200:
            print('Ошибка связи с сайтом. Запустите программу ещё раз')
            print(f'Request url {response.request.url} result {response.status_code}')
            continue

        page = response.text
        soup = BeautifulSoup(page, 'html.parser')
        block = soup.find('div', class_='cardWrapper outerWrapper')
        block2 = block.find('div', class_='wrapper')

        # block3 = block2.find('div', class_='first-row-active-documents')
        for att in block2.findAll('div', class_='attachment'):
            att2 = att.find('span', class_='section__value')
            att3 = att2.find('a')
            src = att3['href']
            file_name = att3['title'].strip()
            dest = os.path.join(lot_tag, file_name)
            dest = make_do_not_exists(dest)

            print(f'Скачиваем файл из {src}')
            file = requests.get(
                src, 
                headers=headers,
                timeout=60)
            print(f'Записываем файл в {dest}')
            with open(dest, 'wb') as f:
                f.write(file.content)

            print(f'Сохранили файл {file_name} из {src} в {dest}')

        record['stage3'] = 'done'
        lots[index] = record

def do_stage_two(wb):
    lots = get_wrapper(wb, 'lots', 2)
    lots_count = len(lots)
    if lots_count <= 0:
        print('В файле не найдено данных о торгах')
        return

    common_info = 'https://zakupki.gov.ru/epz/order/notice/ea44/view/common-info.html'
    supp_base = 'https://zakupki.gov.ru/epz/order/notice/ea44/view/supplier-results.html'
    
    for index in range(lots_count):
        record = lots[index]
        if record['stage2'] == 'done':
            continue

        print(f'Аукцион #{index+1:3d} из {lots_count:3d}, на сумму {record["price"]} руб., опубликован {record["published"]}, название `{record["name"]}`')
        
        if record['stage'] == 'Определение поставщика отменено':
            print(f'{record["stage"]}, пропускаю.')
            record['stage2'] = 'done'
            lots[index] = record
            continue
 
        number = record['number'].replace('"', '')
        params = {'regNumber': number}

        # скачиваем инфу о поставщике
        try:
            response = requests.get(
            supp_base,
            headers=headers,
            params=params,
            timeout=60)
        except Exception as e:
            print(f'Какая-то ошибка с получением информации с сайта, запустите программу ещё раз на второй этап')
            print('Вот описание ошибки:')
            print(e)
            record['stage2'] = 'error'
            lots[index] = record
            continue
            
        if response.status_code != 200:
            print('Ошибка связи с сайтом. Запустите программу ещё раз')
            print(f'Request url {response.request.url} result {response.status_code}')
            record['stage2'] = 'error'
            lots[index] = record
            continue
         
        page = response.text
        soup = BeautifulSoup(page, 'html.parser')
        block = soup.find('div', class_='cardWrapper outerWrapper')
        block2 = block.find('div', class_='wrapper')
        block3 = block2.find('div', class_='cardHeaderBlock')
        block3_2 = block3.find_next_sibling('div')
        block4 = block3_2.find('div', class_='row blockInfo')
        
        for table in block4.find_all('table'):
            thead = table.find('thead')
            thead_tr = thead.find('tr')
            thead_tr_th = thead_tr.find('th')
            caption = thead_tr_th.text.strip()
            if caption.startswith('Заказчик'):
                tbody = table.find('tbody')
                td = tbody.find('td')
                record['customer'] = td.text.strip()
            elif caption.startswith('Участник'):
                tbody = table.find('tbody')
                i = 0
                for tr in tbody.find_all('tr'):
                    i += 1
                    td = tr.find('td')
                    record[f'supplier{i}_name'] = td.text.strip()
                    td2 = td.find_next_sibling('td')
                    record[f'supplier{i}_status'] = td2.text.strip()
                    td3 = td2.find_next_sibling('td')
                    record[f'supplier{i}_price'] = td3.text.strip()

        # скачиваем инфу о товарах
        
        try:
            response = requests.get(
            common_info,
            headers=headers,
            params=params,
            timeout=60)
        except Exception as e:
            print(f'Какая-то ошибка с получением информации с сайта, запустите программу ещё раз на второй этап')
            print('Вот описание ошибки:')
            print(e)
            record['stage2'] = 'error'
            lots[index] = record
            continue
            
        if response.status_code != 200:
            print('Ошибка связи с сайтом. Запустите программу ещё раз')
            print(f'Request url {response.request.url} result {response.status_code}')
            record['stage2'] = 'error'
            lots[index] = record
            continue
        
        page = response.text
        soup = BeautifulSoup(page, 'html.parser')
        block = soup.find('div', id = 'positionKTRU')
        table = block.find('table')
        thead = table.find('thead')
        tbody = table.find('tbody')

        code_table = {}
        for i, th in enumerate(thead.find_all(['th', 'td'])):
            name = th.text.strip()
            if name.startswith('Код'):
                code_table[i] = 'KTRU'
            elif name.startswith('Наименование'):
                code_table[i] = 'name'
            elif name.startswith('Количество'):
                code_table[i] = 'count'
            elif name.startswith('Цена'):
                code_table[i] = 'price'
            elif name.startswith('Стоимость'):
                code_table[i] = 'value'

        good_number = 0
        for row in tbody.find_all('tr', class_='tableBlock__row'):
##            td1 = row.find('td')
##            cl1 = td1['class']
##            if len(cl1) > 1:
##                assert 'header' in cl1[1]
##                continue
            good_number += 1
            for i, val in enumerate(row.find_all('td')):
                value = val.text.strip().replace('\r', '').replace('\n', '')
                value = value.replace('\xa0', '')
                value = value.replace('\u2264', '<=')
                value = value.replace('\u2265', '>=')
                value = value.replace('\u2070', '0')
                while '  ' in value:
                    value = value.replace('  ', ' ')

                if i in code_table:
                    fancy_name = f'good_{good_number:02d}_{code_table[i]}'
                    record[fancy_name] = value

        record['stage2'] = 'done'
        lots[index] = record
        
if __name__ == "__main__":
    main()
